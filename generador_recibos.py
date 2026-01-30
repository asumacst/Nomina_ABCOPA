"""
Generador de recibos de pago por empleado a partir de la nómina quincenal.
Usa la plantilla Excel 'Plantilla recibos de pago.xlsx' y genera un archivo
por cada empleado en una carpeta (ej: datos/recibos_quincena_YYYYMMDD).
"""

import os
import re
import sys
from copy import copy
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Usar la misma carpeta de datos que el sistema principal
if getattr(sys, "frozen", False):
    _BASE_DIR = os.path.dirname(sys.executable)
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(_BASE_DIR, "datos")
DEFAULT_PLANTILLA = os.path.join(DATA_DIR, "Plantilla recibos de pago.xlsx")


def _normalize_col(name):
    """Normaliza nombre de columna para comparación (sin tildes)."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    n = str(name).strip()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ü", "u")]:
        n = n.replace(a, b).replace(a.upper(), b.upper())
    return n


def _find_column(df, *candidates):
    """Devuelve el nombre real de la columna en df que coincida con alguno de los candidatos (normalizado)."""
    norm_df = {_normalize_col(c): c for c in df.columns}
    for cand in candidates:
        cn = _normalize_col(cand)
        if cn in norm_df:
            return norm_df[cn]
    return None


def _safe_value(row, col_name, default=0):
    if col_name is None or col_name not in row.index:
        return default
    v = row[col_name]
    if pd.isna(v):
        return default
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    return v


def _safe_str(v, default=""):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    return str(v).strip() or default


def _nombre_archivo_seguro(nombre, id_emp):
    """Genera un nombre de archivo seguro a partir de nombre e ID."""
    s = f"{_safe_str(nombre)}_{_safe_str(id_emp)}"
    s = re.sub(r'[<>:"/\\|?*]', "_", s)
    s = s.strip("._ ") or "recibo"
    return s[:80]


def generar_recibos(
    archivo_nomina,
    plantilla=None,
    carpeta_salida=None,
    incluir_seguridad=False,
):
    """
    Genera un recibo de pago (Excel) por cada empleado de la nómina, usando la plantilla.

    Args:
        archivo_nomina: Ruta al Excel de nómina quincenal (ej: nomina_quincenal_pago_YYYYMMDD.xlsx).
        plantilla: Ruta a la plantilla Excel. Si es None, usa datos/Plantilla recibos de pago.xlsx.
        carpeta_salida: Carpeta donde guardar los recibos. Si es None, se crea datos/recibos_quincena_YYYYMMDD.
        incluir_seguridad: Si True, incluye también empleados de nómina de seguridad (*_seguridad.xlsx).

    Returns:
        (carpeta_salida, cantidad_recibos) o (None, 0) si hay error.
    """
    if plantilla is None:
        plantilla = DEFAULT_PLANTILLA
    if not os.path.exists(plantilla):
        print(f"[ERROR] No se encontró la plantilla: {plantilla}")
        return None, 0
    if not os.path.exists(archivo_nomina):
        print(f"[ERROR] No se encontró el archivo de nómina: {archivo_nomina}")
        return None, 0

    try:
        df = pd.read_excel(archivo_nomina)
    except Exception as e:
        print(f"[ERROR] No se pudo leer la nómina: {e}")
        return None, 0

    if df.empty:
        print("[ERROR] El archivo de nómina no tiene filas.")
        return None, 0

    # Resolver carpeta de salida
    if carpeta_salida is None:
        # Extraer fecha del nombre del archivo (nomina_quincenal_pago_20260131.xlsx) o de la primera fila
        base = os.path.splitext(os.path.basename(archivo_nomina))[0]
        if "202" in base:
            import re as re2
            m = re2.search(r"(\d{8})", base)
            fecha_str = m.group(1) if m else datetime.now().strftime("%Y%m%d")
        else:
            fecha_str = datetime.now().strftime("%Y%m%d")
        carpeta_salida = os.path.join(DATA_DIR, f"recibos_quincena_{fecha_str}")
    os.makedirs(carpeta_salida, exist_ok=True)

    # Nombres de columnas (soportar encoding)
    col_nombre = _find_column(df, "Nombre")
    col_id = _find_column(df, "ID")
    col_cargo = _find_column(df, "Cargo")
    col_salario_base = _find_column(df, "Salario Base")
    col_quincena_inicio = _find_column(df, "Quincena Inicio")
    col_quincena_fin = _find_column(df, "Quincena Fin")
    col_fecha_pago = _find_column(df, "Fecha de Pago")
    col_horas = _find_column(df, "Total Horas Trabajadas")
    col_horas_extra = _find_column(df, "Horas Extra (después 3 PM)")
    col_pago_extra = _find_column(df, "Pago Extra (25% adicional)")
    col_bono = _find_column(df, "Bono Horas Extra")
    col_horas_feriado = _find_column(df, "Horas Feriado/Domingo")
    col_pago_feriado = _find_column(df, "Pago Feriado/Domingo (50% adicional)")
    col_seguro_social = _find_column(df, "Seguro Social (9.75%)")
    col_seguro_educativo = _find_column(df, "Seguro Educativo (1.25%)")
    col_islr = _find_column(df, "ISLR")
    col_descuento_prestamo = _find_column(df, "Descuento Préstamo")
    col_total_descuentos = _find_column(df, "Total Descuentos")
    col_banco = _find_column(df, "Banco")
    col_cuenta = _find_column(df, "Número de Cuenta")
    col_total_pago = _find_column(df, "Total Pago a Empleados")

    if not col_nombre or not col_id:
        print("[ERROR] La nómina debe tener al menos columnas 'Nombre' e 'ID'.")
        return None, 0

    cantidad = 0
    for idx, row in df.iterrows():
        nombre = _safe_str(row.get(col_nombre))
        id_emp = _safe_str(row.get(col_id))
        if not nombre and not id_emp:
            continue

        total_horas = _safe_value(row, col_horas, 0)
        total_pago = _safe_value(row, col_total_pago, 0)
        total_descuentos = _safe_value(row, col_total_descuentos, 0)
        salario_bruto = total_pago + total_descuentos

        if total_horas and total_horas > 0:
            salario_por_hora = round(salario_bruto / total_horas, 2)
        else:
            salario_por_hora = _safe_value(row, col_salario_base, 0)
        salario_diario = round(salario_por_hora * 8, 2)

        horas_extra = _safe_value(row, col_horas_extra, 0)
        horas_feriado = _safe_value(row, col_horas_feriado, 0)
        horas_regulares = max(0, total_horas - horas_extra - horas_feriado)
        pago_regular = salario_bruto - _safe_value(row, col_pago_extra, 0) - _safe_value(row, col_pago_feriado, 0) - _safe_value(row, col_bono, 0)
        if pago_regular < 0:
            pago_regular = round(salario_por_hora * horas_regulares, 2)

        try:
            wb = openpyxl.load_workbook(plantilla)
        except Exception as e:
            print(f"[ERROR] No se pudo cargar la plantilla: {e}")
            return carpeta_salida, cantidad
        sh = wb.active

        # Reemplazar Franspa/FRANSPA por ABCOPA en toda la hoja
        for row_cell in sh.iter_rows(min_row=1, max_row=sh.max_row, min_col=1, max_col=sh.max_column):
            for cell in row_cell:
                if cell.value is not None and isinstance(cell.value, str):
                    v = str(cell.value)
                    if "FRANSPA" in v.upper():
                        cell.value = v.replace("FRANSPA", "ABCOPA").replace("Franspa", "ABCOPA")

        # Título: dos filas centradas (dividir A1:G2 en A1:G1 y A2:G2 si viene fusionado)
        _titulo_fila1 = "INGENIERIA ABCOPA, S.A"
        _titulo_fila2 = "COMPROBANTE DE PAGO"
        _centro = Alignment(horizontal="center", vertical="center")
        # Desfusionar A1:G2 si existe (plantilla antigua)
        for mr in list(sh.merged_cells.ranges):
            if mr.min_row == 1 and mr.max_row == 2 and mr.min_col == 1 and mr.max_col == 7:
                sh.unmerge_cells(str(mr))
                break
        # Asegurar dos bloques fusionados para las dos filas del título
        if not any(m.min_row == 1 and m.max_row == 1 and m.min_col == 1 and m.max_col == 7 for m in sh.merged_cells.ranges):
            sh.merge_cells("A1:G1")
        if not any(m.min_row == 2 and m.max_row == 2 and m.min_col == 1 and m.max_col == 7 for m in sh.merged_cells.ranges):
            sh.merge_cells("A2:G2")
        sh["A1"].value = _titulo_fila1
        sh["A1"].alignment = _centro
        sh["A2"].value = _titulo_fila2
        sh["A2"].alignment = _centro

        # Encabezado empleado: Proyecto siempre QUINTAS DEL ESTE (no el cargo)
        sh["A4"] = nombre
        sh["B4"] = id_emp
        sh["C4"] = "QUINTAS DEL ESTE"
        sh["D4"] = ""
        sh["F4"] = _safe_str(row.get(col_fecha_pago))
        sh["G4"] = total_pago

        # Rango de quincena
        q_inicio = _safe_str(row.get(col_quincena_inicio))
        q_fin = _safe_str(row.get(col_quincena_fin))
        sh["A7"] = f"SALARIO DEL {q_inicio} AL {q_fin}" if (q_inicio and q_fin) else sh["A7"].value

        # Detalles de pago
        sh["C8"] = salario_por_hora
        sh["C9"] = salario_diario
        sh["C10"] = total_horas
        sh["E8"] = _safe_value(row, col_seguro_social, 0)
        sh["E9"] = _safe_value(row, col_seguro_educativo, 0)
        sh["E10"] = _safe_value(row, col_islr, 0)
        sh["E11"] = total_descuentos
        # Descuentos personales: E14 = etiqueta "MONTO"; D15 = banco (acreedor), E15 = vacío (lo llena el administrador); D16 = ABCOPA, E16 = vacío; E17 = total descuentos personales
        sh["E14"] = "MONTO"
        sh["D15"] = _safe_str(row.get(col_banco))
        sh["E15"] = ""  # Descuento por nómina del banco: lo llena el administrador
        sh["D16"] = "ABCOPA"
        sh["E16"] = ""  # Lo llena el administrador si aplica
        descuento_prestamo = _safe_value(row, col_descuento_prestamo, 0)
        sh["E17"] = descuento_prestamo  # Total descuentos personales (junto a la celda TOTAL bajo ABCOPA)

        # Resumen (bruto = neto + descuentos)
        sh["G8"] = salario_bruto
        sh["G9"] = salario_bruto
        sh["G12"] = total_pago
        sh["G20"] = total_pago

        # Detalle horas: B/C por tipo; B17 = total horas, C17 = total a pagar (suma de regulares + extras + domingo + feriado)
        sh["B13"] = round(horas_regulares, 2)
        sh["C13"] = round(pago_regular, 2)
        sh["B14"] = round(horas_extra, 2)
        sh["C14"] = _safe_value(row, col_pago_extra, 0)
        sh["B15"] = round(horas_feriado, 2)
        sh["C15"] = _safe_value(row, col_pago_feriado, 0)
        sh["B16"] = 0
        sh["C16"] = 0
        sh["B17"] = total_horas
        total_pago_detalle_horas = round(pago_regular + _safe_value(row, col_pago_extra, 0) + _safe_value(row, col_pago_feriado, 0), 2)
        sh["C17"] = total_pago_detalle_horas  # Total a pagar sumando regulares + extras + domingo + feriado (al lado del total de horas)
        sh["C17"].font = Font(bold=True)  # Resaltar total en negrita

        # Sección A19: datos de pago. A19/B19/C19 son etiquetas; debajo van los valores
        sh["A19"] = "DEPOSITADO A LA CUENTA"
        sh["B19"] = "POR LA SUMA DE"
        sh["C19"] = "BANCO"
        sh["A20"] = _safe_str(row.get(col_cuenta)) if col_cuenta and col_cuenta in row.index else ""
        sh["B20"] = total_pago
        sh["C20"] = _safe_str(row.get(col_banco))

        # Firma: nombre del empleado centrado en la celda
        celda_nombre = sh["A22"]
        celda_nombre.value = nombre
        celda_nombre.alignment = Alignment(horizontal="center", vertical="center")

        # Ensanchar columnas de la zona de firma (RECIBIDO CONFORME y celdas inferiores) para que el empleado tenga espacio para firmar
        ancho_firma = 28
        for col in ("D", "E"):
            cd = sh.column_dimensions[col]
            cd.width = max(cd.width or 0, ancho_firma)

        # Duplicar todo el recibo más abajo en la misma hoja (copia para empleado y copia para empresa en una sola página al imprimir)
        _filas_recibo = 22
        _offset_copia = 26  # la segunda copia empieza en la fila 27
        for r in range(1, _filas_recibo + 1):
            for c in range(1, 8):
                src_cell = sh.cell(row=r, column=c)
                tgt_cell = sh.cell(row=r + _offset_copia, column=c)
                tgt_cell.value = src_cell.value
                if getattr(src_cell, "font", None):
                    tgt_cell.font = copy(src_cell.font)
                if getattr(src_cell, "alignment", None):
                    tgt_cell.alignment = copy(src_cell.alignment)
                if getattr(src_cell, "border", None):
                    tgt_cell.border = copy(src_cell.border)
                if getattr(src_cell, "fill", None):
                    tgt_cell.fill = copy(src_cell.fill)
                if getattr(src_cell, "number_format", None):
                    tgt_cell.number_format = src_cell.number_format
        # Replicar las celdas fusionadas en el bloque de la copia
        for mr in list(sh.merged_cells.ranges):
            if mr.max_row <= _filas_recibo:
                r1, c1, r2, c2 = mr.min_row + _offset_copia, mr.min_col, mr.max_row + _offset_copia, mr.max_col
                sh.merge_cells(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")

        # Fila 48 (firma de la copia): misma altura que fila 22 para tener espacio para firmar
        _altura_firma = 25
        if 22 in sh.row_dimensions and sh.row_dimensions[22].height is not None:
            _altura_firma = sh.row_dimensions[22].height
        sh.row_dimensions[22].height = _altura_firma
        sh.row_dimensions[48].height = _altura_firma

        nombre_archivo = _nombre_archivo_seguro(nombre, id_emp) + ".xlsx"
        ruta_recibo = os.path.join(carpeta_salida, nombre_archivo)
        try:
            wb.save(ruta_recibo)
            cantidad += 1
        except Exception as e:
            print(f"[ADVERTENCIA] No se pudo guardar recibo para {nombre}: {e}")

    if incluir_seguridad:
        base_nomina = archivo_nomina.rsplit(".", 1)[0] if "." in archivo_nomina else archivo_nomina
        archivo_seguridad = base_nomina + "_seguridad.xlsx"
        if os.path.exists(archivo_seguridad):
            c2, n2 = generar_recibos(
                archivo_seguridad,
                plantilla=plantilla,
                carpeta_salida=carpeta_salida,
                incluir_seguridad=False,
            )
            if c2:
                cantidad += n2

    return carpeta_salida, cantidad


def main():
    """Ejemplo de uso por línea de comandos."""
    import argparse
    parser = argparse.ArgumentParser(description="Generar recibos de pago a partir de la nómina.")
    parser.add_argument("nomina", nargs="?", default=None, help="Archivo Excel de nómina quincenal")
    parser.add_argument("--plantilla", default=None, help="Ruta a la plantilla Excel")
    parser.add_argument("--carpeta", default=None, help="Carpeta de salida para los recibos")
    parser.add_argument("--seguridad", action="store_true", help="Incluir nómina de seguridad")
    args = parser.parse_args()

    if args.nomina:
        archivo = args.nomina
    else:
        # Buscar el último archivo de nómina en datos
        candidatos = [f for f in os.listdir(DATA_DIR) if f.startswith("nomina_quincenal_pago_") and f.endswith(".xlsx") and "_seguridad" not in f]
        if not candidatos:
            print("No se encontró archivo de nómina en datos/. Indique la ruta: generador_recibos.py <archivo_nomina>")
            return
        candidatos.sort(reverse=True)
        archivo = os.path.join(DATA_DIR, candidatos[0])
        print(f"Usando nómina: {archivo}")

    carpeta, n = generar_recibos(
        archivo,
        plantilla=args.plantilla,
        carpeta_salida=args.carpeta,
        incluir_seguridad=args.seguridad,
    )
    if carpeta is not None:
        print(f"[OK] Se generaron {n} recibos en: {carpeta}")


if __name__ == "__main__":
    main()
